"""Inspect xlsx structure for auction result files. Output JSON."""
import sys, os, json, glob
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

candidates = (
    glob.glob(r"F:\mapa_leilao_perolas_*.xlsx")
    + [
        r"F:\Leilao_Perolas_Cachoeirao_Resultado (1).xlsx",
        r"F:\Leilao_JMP_Supreme_Resultado.xlsx",
        r"F:\Leilao_IPB_Prime_Resultado.xlsx",
    ]
)

def serialize(v):
    if v is None: return None
    if isinstance(v, (str, int, float, bool)): return v
    return str(v)

out = {}
for path in candidates:
    if not os.path.exists(path):
        print(f"MISSING: {path}", file=sys.stderr); continue
    key = os.path.basename(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        out[key] = {"error": str(e)}; continue
    sheets = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([serialize(v) for v in row])
        sheets[sh] = rows
    out[key] = sheets

outpath = r"F:\leilao_inspect.json"
with open(outpath, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"Written {outpath}")
