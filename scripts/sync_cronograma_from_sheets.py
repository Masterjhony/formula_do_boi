"""
Sincroniza a tabela cronograma_leiloes no Supabase para ser um ESPELHO EXATO
da planilha "ESCALA LEILÕES 2026" (Google Sheets).

Fonte: https://docs.google.com/spreadsheets/d/1rzEUSB1Rt4DQ7xlj3Wej4Rn-NwnMSgGk

Comportamento (modelo espelho — decisão do operador, 2026-05-21):
  - Detecta cabeçalhos por nome (fuzzy match) — tolera variação de
    layout entre abas.
  - UPSERT em cronograma_leiloes por (data, nome normalizado):
      • Match  → UPDATE só dos campos que a planilha preenche.
      • Sem match → INSERT.
  - ESPELHO: leilões FUTUROS que estão no banco mas não estão mais na
    planilha são APAGADOS (fantasmas de renomeação/remoção). Antes de
    apagar, catálogo/capa do fantasma são transferidos pro leilão
    sobrevivente mais parecido na mesma data.
  - Registros PASSADOS (data < hoje) nunca são tocados — histórico.
  - Guardas anti-catástrofe:
      • aborta se a extração devolver menos de 20 leilões (download/
        layout quebrado);
      • só apaga em meses que a planilha efetivamente produziu linhas —
        se uma aba não foi parseada, aquele mês fica intocado em vez de
        ser esvaziado.

Uso:
    export SUPABASE_SERVICE_ROLE_KEY=...
    python scripts/sync_cronograma_from_sheets.py

Acompanha scripts/upload_leilao_images.py — ambos rodam pela mesma
GitHub Action (.github/workflows/sync-leiloes.yml).
"""

import datetime as _dt
import difflib
import os
import re
import sys
import tempfile

import openpyxl
import requests

# Garante stdout em utf-8 (logs limpos no GitHub Actions e em terminais Windows)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

GSHEETS_ID = "1rzEUSB1Rt4DQ7xlj3Wej4Rn-NwnMSgGk"
GSHEETS_EXPORT = (
    f"https://docs.google.com/spreadsheets/d/{GSHEETS_ID}/export?format=xlsx"
)

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://hghtikjaqixglmpujbwj.supabase.co"
)
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_KEY:
    raise SystemExit("Defina SUPABASE_SERVICE_ROLE_KEY no ambiente")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

# Mínimo de leilões plausível na planilha do ano inteiro. Abaixo disso,
# a extração quase certamente quebrou — abortamos sem mexer no banco.
MIN_ROWS_SANITY = 20


# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

def norm(s):
    if s is None:
        return ""
    s = str(s).lower().strip()
    repl = {
        "ã": "a", "â": "a", "á": "a", "à": "a",
        "é": "e", "ê": "e", "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u", "ç": "c",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def cell(v):
    """Strip + normaliza None/empty pra None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_date(v):
    """Retorna ISO YYYY-MM-DD ou None."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Excel serial
    try:
        f = float(s)
        if f > 30000:
            base = _dt.datetime(1899, 12, 30)
            return (base + _dt.timedelta(days=f)).date().isoformat()
    except (ValueError, TypeError):
        pass
    # YYYY-MM-DD prefix
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m:
        try:
            _dt.date.fromisoformat(m.group(1))
            return m.group(1)
        except ValueError:
            pass
    return None


def parse_hora(v):
    """Normaliza hora pra HH:MM string. None se não der."""
    if v is None:
        return None
    if hasattr(v, "strftime"):  # datetime.time
        try:
            return v.strftime("%H:%M")
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None
    # "09:00:00" → "09:00"
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    # "9.5" (excel time as float fraction of day)
    try:
        f = float(s)
        if 0 <= f < 1:
            total_min = round(f * 24 * 60)
            return f"{total_min // 60:02d}:{total_min % 60:02d}"
        # "19.0" — tratado como hora cheia
        if 0 <= f <= 23:
            return f"{int(f):02d}:00"
    except (ValueError, TypeError):
        pass
    return s[:5] if len(s) >= 4 else None


def parse_int(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


# ────────────────────────────────────────────────────────────────────────────
# Header mapping
# ────────────────────────────────────────────────────────────────────────────

# canonical_field → list of normalized header aliases (em ordem de prioridade)
HEADER_ALIASES = {
    "data":           ["mes", "dia", "data"],          # col com a data ISO
    "dia_semana":     ["diadasemana"],
    "hora":           ["hora"],
    "nome":           ["leilao"],
    "criador":        ["criador"],
    "presencial":     ["presencial"],
    "leiloeira":      ["leiloeira"],
    "raca":           ["raca"],
    "qtd_animais":    ["qtdanimais", "qtd"],
    "sexo":           ["sexo"],
    "comissao":       ["negociacaodecomissao", "comissao"],
    "contrato":       ["contrato"],
}


def detect_columns(ws):
    """
    Lê linhas 1..3 procurando os cabeçalhos. Retorna dict canonical→col(1-based).
    Ignora colunas não mapeadas. Pula 'comissao' header da SEÇÃO (linha 1) que
    se sobrepõe ao alias do campo.
    """
    found = {}
    # Linha 2 é o cabeçalho de campo (MÊS, HORA, LEILÃO...); linha 1 é só o
    # rótulo da seção (INFORMAÇÕES DO LEILÃO, COMISSÃO...). Linha 2 vence.
    header_cells = {}
    for row in (2, 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            n = norm(v)
            if n and col not in header_cells:
                header_cells[col] = n

    # match canonical → col
    for canonical, aliases in HEADER_ALIASES.items():
        for col, header_norm in header_cells.items():
            if col in found.values():
                continue
            for alias in aliases:
                if header_norm == alias:
                    found[canonical] = col
                    break
            if canonical in found:
                break
    return found


# ────────────────────────────────────────────────────────────────────────────
# Supabase client (REST)
# ────────────────────────────────────────────────────────────────────────────

def fetch_existing():
    """Lê tudo de cronograma_leiloes pra match local (evita N round-trips)."""
    out = []
    page = 0
    while True:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/cronograma_leiloes",
            headers={**HEADERS, "Range-Unit": "items", "Range": f"{page * 1000}-{(page + 1) * 1000 - 1}"},
            params={"select": "id,data,nome,dia_semana,hora,criador,presencial,leiloeira,raca,qtd_animais,sexo,comissao,contrato,catalogo_url,img"},
        )
        if r.status_code not in (200, 206):
            raise RuntimeError(f"Falha ao listar cronograma: {r.status_code} {r.text}")
        chunk = r.json()
        out.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    return out


def insert_row(payload):
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/cronograma_leiloes",
        headers={**HEADERS, "Prefer": "return=representation"},
        json=payload,
    )
    if r.status_code not in (200, 201):
        print(f"  ✗ INSERT falhou: {r.status_code} {r.text}", flush=True)
        return None
    d = r.json()
    return d[0] if isinstance(d, list) else d


def update_row(rec_id, payload):
    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/cronograma_leiloes?id=eq.{rec_id}",
        headers={**HEADERS, "Prefer": "return=minimal"},
        json=payload,
    )
    return r.status_code in (200, 204)


def delete_row(rec_id):
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/cronograma_leiloes?id=eq.{rec_id}",
        headers={**HEADERS, "Prefer": "return=minimal"},
    )
    return r.status_code in (200, 204)


# ────────────────────────────────────────────────────────────────────────────
# Pipeline
# ────────────────────────────────────────────────────────────────────────────

def extract_rows_from_sheet(ws):
    """Yields dicts com os campos mapeados. Pula linhas sem (data + nome)."""
    cols = detect_columns(ws)
    if "data" not in cols or "nome" not in cols:
        return
    for r_idx in range(1, ws.max_row + 1):
        data_iso = parse_date(ws.cell(row=r_idx, column=cols["data"]).value)
        if not data_iso:
            continue
        nome = cell(ws.cell(row=r_idx, column=cols["nome"]).value)
        if not nome:
            continue  # linha de calendário sem leilão
        row = {"data": data_iso, "nome": nome}
        for canonical, col in cols.items():
            if canonical in ("data", "nome"):
                continue
            v = ws.cell(row=r_idx, column=col).value
            if canonical == "hora":
                row[canonical] = parse_hora(v)
            elif canonical == "qtd_animais":
                row[canonical] = parse_int(v)
            elif canonical == "dia_semana":
                # Padroniza Capitalização: "domingo" → "Domingo"
                s = cell(v)
                row[canonical] = s.capitalize() if s else None
            else:
                row[canonical] = cell(v)
        yield row


def carry_over_assets(ghost, index, sheet_keys):
    """Antes de apagar um fantasma, transfere catálogo/capa pro leilão
    sobrevivente (que está na planilha) mais parecido na mesma data."""
    cat = ghost.get("catalogo_url")
    img = ghost.get("img")
    if not cat and not img:
        return
    data_iso = ghost["data"][:10]
    survivors = [
        rec for (d, n), rec in index.items()
        if d == data_iso and (d, n) in sheet_keys and rec["id"] != ghost["id"]
    ]
    if not survivors:
        return
    best, best_score = None, 0.0
    for s in survivors:
        score = difflib.SequenceMatcher(
            None, norm(ghost["nome"]), norm(s["nome"])
        ).ratio()
        if score > best_score:
            best, best_score = s, score
    if not best or best_score < 0.4:
        return
    patch = {}
    if cat and not best.get("catalogo_url"):
        patch["catalogo_url"] = cat
    if img and not best.get("img"):
        patch["img"] = img
    if patch and update_row(best["id"], patch):
        best.update(patch)
        print(f"     ↪ anexos ({', '.join(patch)}) transferidos para '{best['nome']}'", flush=True)


def main():
    print(f"📥 Baixando planilha {GSHEETS_ID}...", flush=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name
    try:
        # Retry: o export do Google Sheets ocasionalmente devolve resposta
        # incompleta (chunked encoding). Tentamos até 3x antes de falhar.
        last_err = None
        for attempt in range(3):
            try:
                r = requests.get(GSHEETS_EXPORT, allow_redirects=True, timeout=60)
                r.raise_for_status()
                open(xlsx_path, "wb").write(r.content)
                last_err = None
                break
            except (requests.exceptions.ChunkedEncodingError,
                    requests.exceptions.ConnectionError,
                    requests.exceptions.Timeout) as e:
                last_err = e
                print(f"  ⚠ tentativa {attempt + 1} falhou: {e}", flush=True)
        if last_err:
            raise last_err

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        all_rows = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            count_before = len(all_rows)
            for row in extract_rows_from_sheet(ws):
                all_rows.append(row)
            print(f"  • {sn}: {len(all_rows) - count_before} leilões", flush=True)

        print(f"\n📊 Total extraído: {len(all_rows)} leilões", flush=True)

        # Guarda anti-catástrofe: a planilha tem 100+ leilões no ano. Se a
        # extração devolveu quase nada, o download/layout quebrou — aborta
        # antes que o espelho apague o banco inteiro.
        if len(all_rows) < MIN_ROWS_SANITY:
            raise SystemExit(
                f"❌ Extração devolveu só {len(all_rows)} leilões "
                f"(mínimo esperado: {MIN_ROWS_SANITY}). Download/layout "
                "provavelmente quebrado — abortando SEM sincronizar."
            )

        # Chave de identidade do leilão na planilha + meses cobertos.
        sheet_keys = {(row["data"], norm(row["nome"])) for row in all_rows}
        sheet_months = {row["data"][:7] for row in all_rows}  # {"2026-06", ...}

        existing = fetch_existing()
        # index: (data, nome_norm) → record
        index = {}
        for rec in existing:
            key = (rec["data"][:10], norm(rec["nome"]))
            index[key] = rec

        # ── UPSERT — adiciona/atualiza tudo que está na planilha ──────────
        inserted = updated = unchanged = 0
        for row in all_rows:
            key = (row["data"], norm(row["nome"]))
            existing_rec = index.get(key)

            # Filtra payload — só envia campos que a planilha preencheu
            payload = {k: v for k, v in row.items() if v not in (None, "")}

            if existing_rec is None:
                rec = insert_row(payload)
                if rec:
                    inserted += 1
                    index[key] = rec
                    print(f"  ➕ {row['data']} | {row['nome']}", flush=True)
            else:
                # Detecta diff — só PATCH se mudou algo
                diff = {}
                for k, v in payload.items():
                    if k == "nome":
                        continue  # nome é a chave; só atualiza outros campos
                    cur = existing_rec.get(k)
                    if cur != v and not (cur in (None, "") and v in (None, "")):
                        diff[k] = v
                if diff:
                    if update_row(existing_rec["id"], diff):
                        updated += 1
                        existing_rec.update(diff)
                        print(f"  🔄 {row['data']} | {row['nome']}  ({', '.join(diff.keys())})", flush=True)
                else:
                    unchanged += 1

        # ── ESPELHO — apaga fantasmas (futuros, em meses que a planilha
        #    cobriu) que não estão mais na planilha ──────────────────────
        today = _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=-3))).date()
        deleted = 0
        for rec in existing:
            data_iso = rec["data"][:10]
            try:
                rec_date = _dt.date.fromisoformat(data_iso)
            except ValueError:
                continue
            if rec_date < today:
                continue  # passado: histórico, nunca apaga
            if data_iso[:7] not in sheet_months:
                continue  # planilha não cobriu esse mês — não arrisca apagar
            key = (data_iso, norm(rec["nome"]))
            if key in sheet_keys:
                continue  # está na planilha — mantém
            # fantasma → preserva catálogo/capa antes de apagar
            carry_over_assets(rec, index, sheet_keys)
            if delete_row(rec["id"]):
                deleted += 1
                print(f"  🗑  {data_iso} | {rec['nome']}  (fora da planilha)", flush=True)

        print(
            f"\n✅ Done — {inserted} inseridos, {updated} atualizados, "
            f"{deleted} removidos, {unchanged} sem mudança ({len(all_rows)} na planilha)",
            flush=True,
        )
    finally:
        try:
            os.unlink(xlsx_path)
        except OSError:
            pass


if __name__ == "__main__":
    sys.exit(main() or 0)
