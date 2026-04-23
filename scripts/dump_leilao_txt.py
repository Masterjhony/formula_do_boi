"""Dump xlsx sheets as readable text per file."""
import sys, os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

files = [
    r"F:\Leilao_Perolas_Cachoeirao_Resultado (1).xlsx",
    r"F:\Leilao_JMP_Supreme_Resultado.xlsx",
    r"F:\Leilao_IPB_Prime_Resultado.xlsx",
    r"F:\mapa_leilao_perolas_cachoeirão.xlsx",
]

for path in files:
    base = os.path.basename(path).replace(' ', '_').replace('(', '').replace(')', '').replace('.xlsx', '')
    out = f"F:/_leilao_{base}.txt"
    wb = openpyxl.load_workbook(path, data_only=True)
    with open(out, 'w', encoding='utf-8') as f:
        for sh in wb.sheetnames:
            ws = wb[sh]
            f.write(f"\n===== SHEET: {sh} ({ws.max_row}x{ws.max_column}) =====\n")
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(v) if v is not None else '' for v in row]
                f.write(f"R{i}: | " + " | ".join(cells) + "\n")
    print(out)
