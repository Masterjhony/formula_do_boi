"""
Sincroniza TODAS as capas embutidas na planilha ESCALA LEILÕES 2026 (fonte:
Google Sheets) para a tabela bula_leiloes.img no Supabase.

Fonte: https://docs.google.com/spreadsheets/d/1rzEUSB1Rt4DQ7xlj3Wej4Rn-NwnMSgGk
       (export automático em formato xlsx)

Para cada imagem embutida nas planilhas mensais:
  1. Identifica (data, leilão, criador) pela linha-âncora
  2. Procura registro em bula_leiloes pela data + similaridade de nome
  3. Se não existe, cria um registro novo
  4. Faz upload ao bucket leilao-covers e atualiza img

Nunca apaga img de registros não-mapeados (o script antigo fazia isso e
apagava capas inseridas manualmente).

Uso:
    export SUPABASE_SERVICE_ROLE_KEY=...
    python scripts/upload_leilao_images.py
"""

import difflib
import os
import re
import sys
import tempfile
import zipfile

import openpyxl
import requests

# Logs sem buffering — no GitHub Actions o stdout é um pipe (fully buffered)
# e as mensagens só apareceriam todas juntas no fim, escondendo em qual passo
# o script estava quando algo deu errado.
try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass

GSHEETS_ID = "1rzEUSB1Rt4DQ7xlj3Wej4Rn-NwnMSgGk"
GSHEETS_EXPORT = (
    f"https://docs.google.com/spreadsheets/d/{GSHEETS_ID}/export?format=xlsx"
)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://hghtikjaqixglmpujbwj.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
BUCKET = "leilao-covers"

if not SUPABASE_KEY:
    raise SystemExit("Defina SUPABASE_SERVICE_ROLE_KEY no ambiente")

HEADERS = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
HEADERS_JSON = {**HEADERS, "Content-Type": "application/json"}


def norm(s):
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"[^a-zà-ÿ0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    repl = {
        "ã": "a", "â": "a", "á": "a", "à": "a",
        "é": "e", "ê": "e", "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u", "ç": "c",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def similar(a, b):
    return difflib.SequenceMatcher(None, norm(a), norm(b)).ratio()


def find_match(records, data_iso, leilao_name, criador_name, used_ids):
    cands = [r for r in records if r["data"][:10] == data_iso and r["id"] not in used_ids]
    if not cands:
        return None
    best, best_score = None, 0.0
    for r in cands:
        score = max(
            similar(r["nome"], criador_name or ""),
            similar(r["nome"], leilao_name or ""),
            similar(r["nome"], f"{leilao_name or ''} {criador_name or ''}"),
        )
        if score > best_score:
            best, best_score = r, score
    if len(cands) == 1 or best_score >= 0.35:
        return best
    return None


def iso_data(v, sheet_name=None):
    import datetime as _dt
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip() if v is not None else ""
    # Excel serial date (e.g., "46141.0")
    try:
        f = float(s)
        if f > 30000:
            base = _dt.datetime(1899, 12, 30)
            return (base + _dt.timedelta(days=f)).date().isoformat()
    except (ValueError, TypeError):
        pass
    # "DD / MM" form — infer year from sheet name (e.g., ABRIL2026)
    parts = s.replace(" ", "").split("/")
    if len(parts) == 2:
        try:
            day, mon = int(parts[0]), int(parts[1])
            year = 2026
            if sheet_name:
                m = re.search(r"(\d{4})", sheet_name)
                if m:
                    year = int(m.group(1))
            return _dt.date(year, mon, day).isoformat()
        except ValueError:
            pass
    return s[:10]


def pretty_nome(criador, leilao):
    src = (criador or leilao or "").strip()
    if not src:
        return "LEILÃO"
    return src.title() if src.isupper() else src


def query_all():
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/bula_leiloes?select=id,nome,data,img&order=data",
        headers=HEADERS_JSON,
    )
    r.raise_for_status()
    return r.json()


def create_record(payload):
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/bula_leiloes",
        headers={**HEADERS_JSON, "Prefer": "return=representation"},
        json=payload,
    )
    if r.status_code not in (200, 201):
        print(f"ERRO criar: {r.status_code} {r.text}")
        return None
    d = r.json()
    return d[0] if isinstance(d, list) else d


def upload_image(rec_id, media_name, data_bytes):
    ext = media_name.rsplit(".", 1)[-1].lower()
    content_type = "image/png" if ext == "png" else "image/jpeg"
    path = f"bula_{rec_id}.{ext}"
    requests.delete(f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{path}", headers=HEADERS)
    r = requests.post(
        f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{path}",
        headers={**HEADERS, "Content-Type": content_type},
        data=data_bytes,
    )
    r.raise_for_status()
    return f"{SUPABASE_URL}/storage/v1/object/public/{BUCKET}/{path}"


def patch_img(rec_id, url):
    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/bula_leiloes?id=eq.{rec_id}",
        headers=HEADERS_JSON,
        json={"img": url},
    )
    return r.status_code in (200, 204)


def download_xlsx(dst):
    """Baixa o xlsx do Google Sheets com retry + validação.

    O export do Google Sheets falha de duas formas: (1) encerra a resposta
    no meio (ChunkedEncodingError) e (2) sob throttle devolve um xlsx
    "só valores", um zip válido mas sem as imagens embutidas — o que faria
    o sync reportar "0 capas" em silêncio, sem erro. Tentamos até 3x e só
    aceitamos o arquivo se ele tiver xl/media/ (as imagens). Se nenhuma
    tentativa trouxer mídia, levanta erro pra falhar o step de forma
    visível, em vez de "passar" sem anexar nada.
    """
    last_err = None
    for attempt in range(3):
        try:
            r = requests.get(GSHEETS_EXPORT, allow_redirects=True, timeout=120)
            r.raise_for_status()
            with open(dst, "wb") as f:
                f.write(r.content)
            with zipfile.ZipFile(dst) as z:
                media = [n for n in z.namelist() if n.startswith("xl/media/")]
            if not media:
                raise RuntimeError(
                    f"xlsx sem xl/media/ ({os.path.getsize(dst)} bytes) — "
                    "resposta truncada/throttled do Google Sheets"
                )
            print(
                f"Download OK: {os.path.getsize(dst)} bytes, "
                f"{len(media)} arquivos de mídia"
            )
            return
        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                zipfile.BadZipFile,
                RuntimeError) as e:
            last_err = e
            print(f"  tentativa {attempt + 1} falhou: {e}")
    raise last_err


def gather_image_map(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        imgs = getattr(ws, "_images", [])
        if not imgs:
            continue
        for idx, img in enumerate(imgs):
            f = img.anchor._from
            row_1b = f.row + 1
            data_val = iso_data(ws.cell(row=row_1b, column=1).value, sn)
            leilao = ws.cell(row=row_1b, column=5).value or ws.cell(row=row_1b + 1, column=5).value
            criador = ws.cell(row=row_1b, column=6).value or ws.cell(row=row_1b + 1, column=6).value
            try:
                raw = img._data() if callable(getattr(img, "_data", None)) else None
            except Exception:
                raw = None
            ext = (getattr(img, "format", None) or "jpg").lower()
            if ext == "jpeg":
                ext = "jpg"
            out.append({
                "sheet": sn,
                "idx": idx,
                "media_name": f"{sn}_img{idx}.{ext}",
                "row_1b": row_1b,
                "data_iso": data_val,
                "leilao": leilao,
                "criador": criador,
                "bytes": raw,
                "ext": ext,
            })
    return out


def main():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        xlsx_path = tmp.name
    try:
        print(f"Baixando do Google Sheets: {GSHEETS_ID}")
        download_xlsx(xlsx_path)

        items = gather_image_map(xlsx_path)
        print(f"Encontradas {len(items)} imagens embutidas\n")

        records = query_all()
        print(f"bula_leiloes: {len(records)} registros\n")

        used_ids = set()
        for it in items:
            if not it["bytes"]:
                print(f"SKIP {it['sheet']} img#{it['idx']}: sem bytes")
                continue
            match = find_match(records, it["data_iso"], it["leilao"], it["criador"], used_ids)
            if match is None:
                payload = {
                    "nome": pretty_nome(it["criador"], it["leilao"]),
                    "data": it["data_iso"],
                    "tipo": "Fêmeas P.O.",
                    "local": "",
                    "horario": "",
                    "modelo": "VIRTUAL",
                    "leiloeira": "BULA",
                    "status": "confirmado",
                }
                match = create_record(payload)
                if not match:
                    continue
                records.append(match)
                print(f"Criado: {match['nome']} ({match['data'][:10]}) id={match['id']}")

            url = upload_image(match["id"], it["media_name"], it["bytes"])
            ok = patch_img(match["id"], url)
            used_ids.add(match["id"])
            tag = "OK " if ok else "ERR"
            print(f"{tag} [{it['sheet']} img#{it['idx']}] {it['data_iso']} | {it['criador']}  ->  '{match['nome']}'")

        print(f"\nDone — {len(used_ids)} registros com capa atualizada.")
    finally:
        try:
            os.unlink(xlsx_path)
        except OSError:
            pass


if __name__ == "__main__":
    sys.exit(main() or 0)
